"""Excel (.xlsx) support for the existing CSV importers (v2.0).

Rather than a second importer, this converts a workbook to CSV text and hands it
to the CSV path — so the column-alias tolerance, the rating rules and the
idempotent hashing are inherited rather than reimplemented, and an .xlsx and a
.csv of the same data import identically.
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime

XLSX_MAGIC = b"PK\x03\x04"


def looks_like_xlsx(data: bytes, filename: str | None = None) -> bool:
    """Content first, filename second: uploads are routinely misnamed, and the
    zip magic is what actually decides whether openpyxl can read it."""
    if data[:4] == XLSX_MAGIC:
        return True
    return bool(filename and filename.lower().endswith((".xlsx", ".xlsm")))


def xlsx_to_csv(data: bytes, sheet: str | None = None) -> str:
    """First worksheet (or `sheet`) as CSV text.

    read_only + values_only so a large export streams instead of building a cell
    object per value, and formulas arrive as their computed result.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as err:  # pragma: no cover - dependency is declared
        raise RuntimeError(
            "Excel import needs the openpyxl package") from err

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as err:
        raise ValueError(
            f"could not read this file as Excel ({type(err).__name__}). "
            "If it is a .xls or an exported HTML table, save it as .xlsx or .csv."
        ) from err

    try:
        ws = wb[sheet] if sheet else wb.worksheets[0]
        if ws is None:
            raise ValueError("the workbook has no worksheets")

        out = io.StringIO()
        writer = csv.writer(out, lineterminator="\n")
        wrote_header = False
        for row in ws.iter_rows(values_only=True):
            cells = [_cell(v) for v in row]
            if not wrote_header:
                # Leading blank/decorative rows are common in exports; the real
                # header is the first row with content. Writing a blank one makes
                # DictReader key every column to '' and match no alias.
                if not any(c.strip() for c in cells):
                    continue
                wrote_header = True
            writer.writerow(cells)
        return out.getvalue()
    finally:
        wb.close()


def _cell(value) -> str:
    """One cell as text, in a form the CSV date/number parsers already accept."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        # Excel stores every number as a float, so a 3-star rating arrives as
        # 3.0 and an id as 40123.0. int() keeps ids and ratings recognisable.
        return str(int(value))
    return str(value)
